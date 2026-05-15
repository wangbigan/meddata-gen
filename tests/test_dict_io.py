"""dict_io 模块单元测试。

覆盖:
    - schemas      元数据查询
    - template_builder  Excel 模板生成与结构校验
    - validators   单元格解析与行级校验
    - importer     dry-run 导入流程与报告
    - builtin_loader  内置数据构建(不连库)
"""
from __future__ import annotations

import os
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from meddata_gen.dict_io import DICT_TABLES, build_template, get_table, tables_for_database
from meddata_gen.dict_io.builtin_loader import (
    _build_antibiotic_rows,
    _build_diagnosis_rows,
    _build_exam_rows,
    _build_lab_rows,
    _build_organism_rows,
    load_builtin_dicts,
)
from meddata_gen.dict_io.importer import ImportReport, import_dicts
from meddata_gen.dict_io.schemas import DictColumn
from meddata_gen.dict_io.validators import parse_row, parse_value, primary_key_index


# ---------- schemas ----------

def test_dict_tables_count():
    assert len(DICT_TABLES) == 8


def test_get_table_existing():
    t = get_table("diagnosis_dict")
    assert t is not None
    assert t.cn_name == "诊断字典"
    assert t.database == "his_db"


def test_get_table_missing():
    assert get_table("not_exist") is None


def test_tables_for_database():
    assert len(tables_for_database("his_db")) == 4
    assert len(tables_for_database("lis_db")) == 3
    assert len(tables_for_database("ris_db")) == 1


# ---------- template_builder ----------

def test_build_template_all_sheets(tmp_path: Path):
    out = tmp_path / "dict.xlsx"
    path = build_template(out, with_samples=True)
    assert path.exists()

    wb = openpyxl.load_workbook(path)
    assert "_README" in wb.sheetnames
    assert len(wb.sheetnames) == 9  # _README + 8 dicts

    # diagnosis_dict sheet 结构
    ws = wb["diagnosis_dict"]
    assert ws.max_row >= 4  # 3 header + at least 1 sample
    assert ws.cell(row=1, column=1).value == "诊断编码*"
    assert ws.cell(row=2, column=1).value == "icd_code"


def test_build_template_no_samples(tmp_path: Path):
    out = tmp_path / "dict_blank.xlsx"
    build_template(out, with_samples=False)
    wb = openpyxl.load_workbook(out)
    ws = wb["diagnosis_dict"]
    assert ws.max_row == 3  # headers only


def test_build_template_subset(tmp_path: Path):
    out = tmp_path / "dict_sub.xlsx"
    build_template(out, include=["lab_items_dict", "organism_dict"], with_samples=False)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["_README", "lab_items_dict", "organism_dict"]


# ---------- validators ----------

def test_parse_value_text():
    col = DictColumn("name", "名称", "text")
    val, err = parse_value(col, "  hello  ")
    assert val == "hello"
    assert err is None


def test_parse_value_int():
    col = DictColumn("age", "年龄", "int")
    assert parse_value(col, 42)[0] == 42
    assert parse_value(col, "42")[0] == 42
    assert parse_value(col, "abc")[1] is not None


def test_parse_value_decimal():
    col = DictColumn("price", "价格", "decimal")
    assert parse_value(col, "3.14")[0] == pytest.approx(3.14)
    assert parse_value(col, "abc")[1] is not None


def test_parse_value_bool():
    col = DictColumn("flag", "标志", "bool")
    assert parse_value(col, True)[0] is True
    assert parse_value(col, "TRUE")[0] is True
    assert parse_value(col, "false")[0] is False
    assert parse_value(col, "0")[0] is False
    assert parse_value(col, "maybe")[1] is not None


def test_parse_value_required():
    col = DictColumn("code", "编码", "text", required=True)
    assert parse_value(col, "")[1] is not None
    assert parse_value(col, None)[1] is not None


def test_parse_value_enum():
    col = DictColumn("type", "类型", "text", enum_values=("A", "B"))
    assert parse_value(col, "A")[1] is None
    assert parse_value(col, "C")[1] is not None


def test_parse_row():
    table = get_table("diagnosis_dict")
    assert table is not None
    raw = {
        "icd_code": "J18.901",
        "diagnosis_name": "肺炎",
        "category": "呼吸系统",
        "is_chronic": "FALSE",
        "is_infectious": "TRUE",
    }
    cleaned, errs = parse_row(table, raw)
    assert len(errs) == 0
    assert cleaned[0] == "J18.901"
    assert cleaned[3] is False
    assert cleaned[4] is True


def test_primary_key_index():
    table = get_table("diagnosis_dict")
    assert primary_key_index(table) == 0


# ---------- importer (dry-run) ----------

def test_import_dry_run(tmp_path: Path):
    template = tmp_path / "dict.xlsx"
    build_template(template, with_samples=True)
    report = import_dicts(str(template), dry_run=True)
    assert isinstance(report, ImportReport)
    assert report.dry_run is True
    total_ok = sum(r.rows_ok for r in report.results)
    assert total_ok > 0


def test_import_with_bad_value(tmp_path: Path):
    template = tmp_path / "dict.xlsx"
    build_template(template, with_samples=True)
    wb = openpyxl.load_workbook(template)
    ws = wb["diagnosis_dict"]
    ws.cell(row=5, column=4, value="bad_bool")
    wb.save(template)

    report = import_dicts(str(template), dry_run=True)
    diag = next(r for r in report.results if r.table_name == "diagnosis_dict")
    assert diag.rows_failed > 0
    assert any("TRUE/FALSE" in e for e in diag.errors)


# ---------- builtin_loader (data building only) ----------

def test_build_diagnosis_rows():
    rows = _build_diagnosis_rows()
    assert len(rows) > 0
    # ICD10_DIAGNOSES 约 120 条
    assert all(len(r) == 5 for r in rows)
    assert any(r[3] is True for r in rows)  # 至少有一个慢性病


def test_build_lab_rows():
    rows = _build_lab_rows()
    assert len(rows) > 0
    assert all(len(r) == 9 for r in rows)


def test_build_organism_rows():
    rows = _build_organism_rows()
    assert len(rows) > 0
    assert all(len(r) == 4 for r in rows)


def test_build_antibiotic_rows():
    rows = _build_antibiotic_rows()
    assert len(rows) > 0


def test_build_exam_rows():
    rows = _build_exam_rows()
    assert len(rows) > 0
    assert all(len(r) == 7 for r in rows)


def test_load_builtin_dicts_mock():
    """mock 数据库连接,验证 load_builtin_dicts 会遍历 8 张表。"""
    mock_conn = MagicMock()
    mock_cur = MagicMock()
    mock_conn.cursor.return_value = mock_cur

    with patch("meddata_gen.dict_io.builtin_loader._connect", return_value=mock_conn):
        stats = load_builtin_dicts()

    assert len(stats) == 8
    assert all(v >= 0 for v in stats.values())
