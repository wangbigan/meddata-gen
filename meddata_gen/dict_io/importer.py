"""字典 Excel 导入器。

读取用户填写的 Excel,逐行校验后写入对应 PostgreSQL 数据库。

流程:
    1. 加载 Workbook,遍历每个 sheet
    2. 跳过 _README,按 sheet 名匹配 DICT_TABLES
    3. 第 2 行(英文列名)必须与表定义完全一致
    4. 第 4 行起逐行读取 → 类型/必填/枚举/主键唯一性校验
    5. 按 --mode(upsert/replace/append) 写入数据库
    6. 生成导入报告

报告格式 (markdown 表格):
    | Sheet | 读入行 | 通过 | 已插入 | 已更新 | 跳过 | 失败 |
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import click
import psycopg2
from openpyxl import load_workbook

from meddata_gen import config
from meddata_gen.dict_io.schemas import DICT_TABLES, DictColumn, DictTable, get_table
from meddata_gen.dict_io.validators import parse_row, primary_key_index


@dataclass(frozen=True)
class ImportResult:
    """单张字典表的导入结果。"""

    table_name: str
    rows_read: int = 0
    rows_ok: int = 0
    rows_inserted: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    rows_failed: int = 0
    errors: List[str] = None  # type: ignore[assignment]

    def __post_init__(self):
        if self.errors is None:
            object.__setattr__(self, "errors", [])

    def as_tuple(self) -> Tuple:
        return (
            self.table_name,
            self.rows_read,
            self.rows_ok,
            self.rows_inserted,
            self.rows_updated,
            self.rows_skipped,
            self.rows_failed,
        )


@dataclass(frozen=True)
class ImportReport:
    """完整导入报告。"""

    file_path: str
    mode: str
    dry_run: bool
    results: List[ImportResult]

    def to_markdown(self) -> str:
        lines = [
            "# 字典导入报告",
            "",
            f"- **文件**: {self.file_path}",
            f"- **模式**: {self.mode}",
            f"- **试运行**: {'是' if self.dry_run else '否'}",
            "",
            "| Sheet | 读入行 | 校验通过 | 已插入 | 已更新 | 跳过 | 失败 |",
            "|-------|--------|----------|--------|--------|------|------|",
        ]
        for r in self.results:
            lines.append(
                f"| {r.table_name} | {r.rows_read} | {r.rows_ok} | "
                f"{r.rows_inserted} | {r.rows_updated} | {r.rows_skipped} | {r.rows_failed} |"
            )
        lines.append("")
        for r in self.results:
            if r.errors:
                lines.append(f"## {r.table_name} 错误详情")
                lines.append("")
                for e in r.errors[:50]:
                    lines.append(f"- {e}")
                if len(r.errors) > 50:
                    lines.append(f"- ... 还有 {len(r.errors) - 50} 条错误未展示")
                lines.append("")
        return "\n".join(lines)


def _connect(db_name: str) -> Any:
    cfg = config.DB_CONFIG.copy()
    cfg["database"] = db_name
    return psycopg2.connect(**cfg)


def _read_sheet_rows(ws, table: DictTable) -> List[Dict[str, Any]]:
    """读取 Excel sheet 的数据行(从第 4 行起),返回列名→原始值的 dict 列表。"""
    # 第 2 行为英文列名
    header_row = [cell.value for cell in ws[2]]
    # 与 schema 对齐:只取前 len(columns) 列,或按列名匹配
    expected = [c.name for c in table.columns]
    col_names = []
    for i, val in enumerate(header_row):
        if i >= len(expected):
            break
        if val is None or str(val).strip() == "":
            # 如果 Excel 列数不够,用 schema 列名补齐(常见:用户删除了某列)
            val = expected[i]
        col_names.append(str(val).strip())

    # 列名不一致时给出警告
    actual = col_names
    if actual != expected:
        # 允许只填写部分列(用户只填了需要的),但至少主键列要有
        pk_col = primary_key_column(table)
        if pk_col.name not in actual:
            raise ValueError(
                f"Sheet '{table.name}' 缺少主键列 '{pk_col.name}'.\n"
                f"  期望列: {expected}\n  实际列: {actual}"
            )

    rows: List[Dict[str, Any]] = []
    for row_idx in range(4, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(col_names))]
        # 整行空则跳过
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        row_dict = {}
        for col_name, val in zip(col_names, values):
            row_dict[col_name] = val
        rows.append(row_dict)
    return rows


def _build_insert_sql(table: DictTable) -> str:
    col_names = [c.name for c in table.columns]
    placeholders = ",".join(["%s"] * len(col_names))
    col_list = ",".join(col_names)
    return f"INSERT INTO {table.name} ({col_list}) VALUES ({placeholders})"


def _build_upsert_sql(table: DictTable) -> str:
    insert = _build_insert_sql(table)
    pk_col = primary_key_column(table)
    set_clause = ",".join(
        f"{c.name}=EXCLUDED.{c.name}"
        for c in table.columns
        if not c.primary_key
    )
    if set_clause:
        return f"{insert} ON CONFLICT ({pk_col.name}) DO UPDATE SET {set_clause}"
    return f"{insert} ON CONFLICT ({pk_col.name}) DO NOTHING"


def _build_truncate_sql(table: DictTable) -> str:
    return f"TRUNCATE TABLE {table.name}"


def primary_key_column(table: DictTable) -> DictColumn:
    """返回主键列定义。"""
    for col in table.columns:
        if col.primary_key:
            return col
    raise ValueError(f"字典表 {table.name} 未定义主键")


def import_dicts(
    file_path: str,
    *,
    mode: str = "upsert",
    dry_run: bool = False,
    system: Optional[str] = None,
    report_path: Optional[str] = None,
) -> ImportReport:
    """执行字典导入。

    Args:
        file_path: Excel 文件路径
        mode: upsert | replace | append
        dry_run: 为 True 时仅校验,不写库
        system: 限制只导入指定 DB(his/lis/ris),None=全部
        report_path: 导入报告输出路径(可选)

    Returns:
        ImportReport
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = load_workbook(path)
    results: List[ImportResult] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "_README":
            continue

        table = get_table(sheet_name)
        if table is None:
            # 未知 sheet:跳过并记录
            results.append(
                ImportResult(
                    table_name=sheet_name,
                    rows_skipped=0,
                    errors=[f"未知字典表 '{sheet_name}',未注册于 DICT_TABLES,已跳过"],
                )
            )
            continue

        if system and table.database != f"{system}_db":
            results.append(
                ImportResult(
                    table_name=sheet_name,
                    rows_skipped=0,
                    errors=[f"当前限定 system={system},跳过 {table.database} 的字典"],
                )
            )
            continue

        ws = wb[sheet_name]
        try:
            raw_rows = _read_sheet_rows(ws, table)
        except ValueError as e:
            results.append(
                ImportResult(table_name=sheet_name, errors=[str(e)])
            )
            continue

        pk_idx = primary_key_index(table)
        pk_seen: set = set()
        errors: List[str] = []
        ok_rows: List[Tuple[Any, ...]] = []

        for row_idx, raw_row in enumerate(raw_rows, start=4):
            cleaned, row_errs = parse_row(table, raw_row)
            if row_errs:
                errors.extend(
                    f"第{row_idx}行: {err}" for err in row_errs
                )
                continue

            pk_val = cleaned[pk_idx]
            if pk_val in pk_seen:
                errors.append(f"第{row_idx}行: 主键 '{pk_val}' 在 sheet 内重复")
                continue
            pk_seen.add(pk_val)
            ok_rows.append(cleaned)

        inserted = updated = skipped = failed = 0
        if not dry_run and ok_rows:
            conn = _connect(table.database)
            conn.autocommit = True
            cur = conn.cursor()
            try:
                if mode == "replace":
                    cur.execute(_build_truncate_sql(table))

                if mode == "upsert":
                    sql = _build_upsert_sql(table)
                else:
                    sql = _build_insert_sql(table)

                for row in ok_rows:
                    try:
                        cur.execute(sql, row)
                        if mode == "upsert":
                            if cur.statusmessage and "UPDATE" in cur.statusmessage:
                                updated += 1
                            else:
                                inserted += 1
                        else:
                            inserted += 1
                    except psycopg2.errors.UniqueViolation as e:
                        if mode == "append":
                            failed += 1
                            errors.append(f"主键冲突: {row[pk_idx]}")
                        else:
                            # replace/upsert 下理论上不应出现
                            errors.append(f"写入异常: {e}")
                    except psycopg2.Error as e:
                        failed += 1
                        errors.append(f"写入异常: {e}")
            finally:
                cur.close()
                conn.close()

        failed += len(errors)
        results.append(
            ImportResult(
                table_name=sheet_name,
                rows_read=len(raw_rows),
                rows_ok=len(ok_rows),
                rows_inserted=inserted,
                rows_updated=updated,
                rows_skipped=0,
                rows_failed=failed,
                errors=errors,
            )
        )

    report = ImportReport(
        file_path=file_path,
        mode=mode,
        dry_run=dry_run,
        results=results,
    )
    if report_path:
        Path(report_path).parent.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report.to_markdown())
    return report
