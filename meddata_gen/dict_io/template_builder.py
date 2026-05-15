"""字典 Excel 模板生成器。

从 ``schemas.DICT_TABLES`` 驱动,生成可供用户填写的 Excel 模板。

Excel 结构:
    _README                综述 + 字典表清单
    <table_name> (8 个)    每张字典表一个 sheet
        第 1 行: 中文表头(必填项标红 + *)
        第 2 行: 英文列名(程序解析依据)
        第 3 行: 字段说明
        第 4 行起: 示例行 / 用户填写区域
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from meddata_gen.dict_io.schemas import DICT_TABLES, DictColumn, DictTable


# ---------- 样式常量 ----------

_TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
_SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
_BODY_FONT = Font(name="微软雅黑", size=10, color="333333")

_HEADER_FILL_REQUIRED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
_HEADER_FILL_OPTIONAL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_ENG_FONT = Font(name="Consolas", size=9, color="595959", italic=True)
_ENG_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_DESC_FONT = Font(name="微软雅黑", size=9, color="808080")
_DESC_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
_SAMPLE_FONT = Font(name="微软雅黑", size=10, color="595959", italic=True)
_SAMPLE_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

_DATA_VALIDATION_LAST_ROW = 1000  # 预留 1000 行的下拉/校验范围


# ---------- 主入口 ----------

def build_template(
    output_path: str | Path,
    *,
    include: Optional[Iterable[str]] = None,
    with_samples: bool = True,
) -> Path:
    """生成字典 Excel 模板。

    Args:
        output_path: Excel 输出路径(.xlsx)
        include: 限定要包含的字典表名;None=全部 8 张
        with_samples: 是否带示例行(默认 True;False 时只保留表头 3 行)

    Returns:
        实际写入的文件路径
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    if include is None:
        tables: List[DictTable] = list(DICT_TABLES)
    else:
        wanted = {name.strip() for name in include}
        tables = [t for t in DICT_TABLES if t.name in wanted]
        missing = wanted - {t.name for t in tables}
        if missing:
            raise ValueError(f"未知字典表: {sorted(missing)},可选: {[t.name for t in DICT_TABLES]}")
        if not tables:
            raise ValueError("未选中任何字典表")

    wb = Workbook()
    # 移除默认 sheet
    default = wb.active
    wb.remove(default)

    _write_readme_sheet(wb, tables, with_samples=with_samples)
    for table in tables:
        _write_table_sheet(wb, table, with_samples=with_samples)

    wb.save(out)
    return out


# ---------- _README sheet ----------

def _write_readme_sheet(wb: Workbook, tables: List[DictTable], *, with_samples: bool) -> None:
    ws = wb.create_sheet("_README", 0)

    ws["A1"] = "meddata-gen 字典数据模板"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A3"] = "使用说明"
    ws["A3"].font = _SECTION_FONT
    instructions = [
        "1. 本模板包含多个字典 Sheet,Sheet 名 = 字典表名,请勿修改 Sheet 名。",
        "2. 每个字典 Sheet 的前 3 行是表头(中文/英文/说明),从第 4 行开始填写数据。",
        "3. 标红表头 + '*' 为必填列;主键列不可重复。",
        "4. 部分列提供下拉(枚举),仅可从下拉选项中选择。",
        "5. 布尔列填写 TRUE 或 FALSE。",
        "6. 仅填写需要的字典 Sheet 即可,未填写的 Sheet 导入时会自动跳过。",
        "7. 填写完成后,执行: meddata-gen dict-import -f <文件路径>",
        "",
        "若要快速演示,可使用内置示例字典:",
        "    meddata-gen dict-import --use-builtin",
    ]
    for i, line in enumerate(instructions, start=4):
        ws.cell(row=i, column=1, value=line).font = _BODY_FONT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    section_start = 4 + len(instructions) + 2
    ws.cell(row=section_start, column=1, value="字典表清单").font = _SECTION_FONT

    header_row = section_start + 1
    headers = ["序号", "Sheet 名 / 表名", "中文名", "目标数据库", "用途说明"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL_OPTIONAL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _THIN_BORDER

    for idx, t in enumerate(tables, start=1):
        row = header_row + idx
        values = [idx, t.name, t.cn_name, t.database, t.description]
        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = _BODY_FONT
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = _THIN_BORDER

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 60
    ws.row_dimensions[1].height = 30


# ---------- 字典 sheet ----------

def _write_table_sheet(wb: Workbook, table: DictTable, *, with_samples: bool) -> None:
    ws = wb.create_sheet(table.name)

    # 第 1 行: 中文表头
    for col_idx, col in enumerate(table.columns, start=1):
        label = col.cn_name + ("*" if col.required else "")
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL_REQUIRED if col.required else _HEADER_FILL_OPTIONAL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _THIN_BORDER

    # 第 2 行: 英文列名(导入解析的依据)
    for col_idx, col in enumerate(table.columns, start=1):
        c = ws.cell(row=2, column=col_idx, value=col.name)
        c.font = _ENG_FONT
        c.fill = _ENG_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _THIN_BORDER

    # 第 3 行: 字段说明
    for col_idx, col in enumerate(table.columns, start=1):
        desc = col.description
        if col.primary_key:
            desc = (desc + " | 主键").strip(" |")
        if col.enum_values:
            desc = (desc + f" | 取值: {', '.join(col.enum_values)}").strip(" |")
        c = ws.cell(row=3, column=col_idx, value=desc)
        c.font = _DESC_FONT
        c.fill = _DESC_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _THIN_BORDER

    # 第 4 行起: 示例行
    if with_samples and table.sample_rows:
        for r_offset, row in enumerate(table.sample_rows):
            for col_idx, val in enumerate(row, start=1):
                c = ws.cell(row=4 + r_offset, column=col_idx, value=val)
                c.font = _SAMPLE_FONT
                c.fill = _SAMPLE_FILL
                c.border = _THIN_BORDER

    # 冻结前 3 行(表头)
    ws.freeze_panes = "A4"

    # 列宽
    for col_idx, col in enumerate(table.columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _column_width_for(col)

    # 行高
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 38

    # 数据有效性(枚举下拉 / 布尔下拉)
    _add_data_validations(ws, table)


def _column_width_for(col: DictColumn) -> int:
    """根据列类型与中文表头长度决定 Excel 列宽。"""
    base = max(12, len(col.cn_name) * 2 + 4)
    if col.excel_type == "text":
        return min(max(base, 18), 32)
    if col.excel_type == "decimal":
        return 14
    if col.excel_type == "int":
        return 12
    if col.excel_type == "bool":
        return 12
    return base


def _add_data_validations(ws, table: DictTable) -> None:
    """为枚举列与布尔列添加 Excel 数据有效性(下拉框)。"""
    for col_idx, col in enumerate(table.columns, start=1):
        formula: Optional[str] = None
        error_msg: Optional[str] = None
        if col.enum_values:
            formula = '"' + ",".join(col.enum_values) + '"'
            error_msg = f"必须从下拉中选择: {', '.join(col.enum_values)}"
        elif col.excel_type == "bool":
            formula = '"TRUE,FALSE"'
            error_msg = "请填写 TRUE 或 FALSE"

        if formula is None:
            continue

        letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = error_msg
        dv.errorTitle = "取值不合法"
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(f"{letter}4:{letter}{_DATA_VALIDATION_LAST_ROW}")
